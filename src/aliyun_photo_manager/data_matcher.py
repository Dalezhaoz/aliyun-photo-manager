from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple


LogFn = Optional[Callable[[str], None]]


@dataclass
class ColumnMapping:
    target_column: str
    source_column: str


@dataclass
class DataMatchOptions:
    target_path: Path
    source_path: Path
    target_key_column: str
    source_key_column: str
    extra_match_mappings: List[ColumnMapping]
    transfer_mappings: List[ColumnMapping]
    output_path: Optional[Path] = None


@dataclass
class DataMatchSummary:
    target_path: Path
    source_path: Path
    output_path: Path
    target_key_column: str
    source_key_column: str
    extra_match_mappings: List[ColumnMapping]
    transfer_mappings: List[ColumnMapping]
    total_rows: int
    matched_rows: int
    unmatched_rows: int
    duplicate_source_keys: int
    ambiguous_rows: int
    target_header_row: int
    source_header_row: int


def _log(logger: LogFn, message: str) -> None:
    if logger is not None:
        logger(message)


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError(
            "缺少依赖 openpyxl，请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return Workbook, load_workbook


def _load_xlrd():
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(
            "缺少依赖 xlrd，请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return xlrd


def _read_sheet_matrix(file_path: Path) -> List[List[str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        _, load_workbook = _load_openpyxl()
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        matrix: List[List[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            matrix.append([_normalize(value) for value in row])
        return matrix
    if suffix == ".xls":
        xlrd = _load_xlrd()
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        matrix: List[List[str]] = []
        for row_index in range(sheet.nrows):
            matrix.append(
                [_normalize(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)]
            )
        return matrix
    raise ValueError("仅支持 `.xlsx` 或 `.xls` 文件。")


def _trim_row(row: List[str]) -> List[str]:
    trimmed = list(row)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def _detect_header_row(matrix: List[List[str]]) -> int:
    best_index = 0
    best_score = -1
    check_limit = min(len(matrix), 10)
    for row_index in range(check_limit):
        row = _trim_row(matrix[row_index])
        non_empty = [cell for cell in row if cell]
        if not non_empty:
            continue
        unique_count = len(set(non_empty))
        score = len(non_empty) * 10 + unique_count
        # 跳过类似“附件1”这种单格标题，尽量识别下面真正的表头行。
        if len(non_empty) == 1:
            score -= 20
        if score > best_score:
            best_score = score
            best_index = row_index
    return best_index


def _extract_headers_and_rows(matrix: List[List[str]]) -> Tuple[int, List[str], List[List[str]]]:
    if not matrix:
        raise ValueError("表格为空，无法读取表头。")
    header_index = _detect_header_row(matrix)
    raw_headers = _trim_row(matrix[header_index])
    headers: List[str] = []
    for index, header in enumerate(raw_headers):
        headers.append(header or f"未命名列{index + 1}")

    rows: List[List[str]] = []
    width = len(headers)
    for raw_row in matrix[header_index + 1 :]:
        row = list(raw_row[:width])
        if len(row) < width:
            row.extend([""] * (width - len(row)))
        row = [_normalize(value) for value in row]
        if any(row):
            rows.append(row)
    return header_index + 1, headers, rows


def list_headers(file_path: Path) -> List[str]:
    _, headers, _ = _extract_headers_and_rows(_read_sheet_matrix(file_path))
    return headers


def _build_output_path(target_path: Path) -> Path:
    return target_path.with_name(f"{target_path.stem}_数据匹配结果.xlsx")


def _header_indexes(headers: List[str]) -> Dict[str, int]:
    return {header: index for index, header in enumerate(headers)}


def _make_key(
    row_values: List[str],
    header_index: Dict[str, int],
    key_column: str,
    extra_mappings: List[ColumnMapping],
    use_target_side: bool,
) -> Tuple[str, ...]:
    values = [_normalize(row_values[header_index[key_column]])]
    for mapping in extra_mappings:
        column_name = mapping.target_column if use_target_side else mapping.source_column
        values.append(_normalize(row_values[header_index[column_name]]))
    return tuple(values)


def run_data_match(options: DataMatchOptions, logger: LogFn = None) -> DataMatchSummary:
    if not options.target_path.exists():
        raise FileNotFoundError(f"未找到目标表：{options.target_path}")
    if not options.source_path.exists():
        raise FileNotFoundError(f"未找到来源表：{options.source_path}")

    target_header_row, target_headers, target_rows = _extract_headers_and_rows(
        _read_sheet_matrix(options.target_path)
    )
    source_header_row, source_headers, source_rows = _extract_headers_and_rows(
        _read_sheet_matrix(options.source_path)
    )

    if options.target_key_column not in target_headers:
        raise ValueError(f"目标表中不存在匹配列：{options.target_key_column}")
    if options.source_key_column not in source_headers:
        raise ValueError(f"来源表中不存在匹配列：{options.source_key_column}")
    for mapping in options.extra_match_mappings:
        if mapping.target_column not in target_headers:
            raise ValueError(f"目标表中不存在附加匹配列：{mapping.target_column}")
        if mapping.source_column not in source_headers:
            raise ValueError(f"来源表中不存在附加匹配列：{mapping.source_column}")
    for mapping in options.transfer_mappings:
        if mapping.source_column not in source_headers:
            raise ValueError(f"来源表中不存在补充列：{mapping.source_column}")

    target_index = _header_indexes(target_headers)
    source_index = _header_indexes(source_headers)

    source_map: Dict[Tuple[str, ...], List[str]] = {}
    duplicate_keys = set()
    for row in source_rows:
        match_key = _make_key(
            row,
            source_index,
            options.source_key_column,
            options.extra_match_mappings,
            use_target_side=False,
        )
        if not match_key[0]:
            continue
        if match_key in source_map:
            duplicate_keys.add(match_key)
            source_map[match_key] = []
        else:
            source_map[match_key] = row

    Workbook, _ = _load_openpyxl()
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "匹配结果"

    output_headers = list(target_headers)
    existing_header_set = set(output_headers)
    final_transfer_headers: List[str] = []
    for mapping in options.transfer_mappings:
        header_name = mapping.target_column or mapping.source_column
        if header_name in existing_header_set:
            header_name = f"{header_name}(匹配)"
        existing_header_set.add(header_name)
        final_transfer_headers.append(header_name)
    output_sheet.append(output_headers + final_transfer_headers)

    report_sheet = output_workbook.create_sheet("匹配结果清单")
    report_sheet.append(["行号", "匹配键", "状态", "备注"])

    total_rows = len(target_rows)
    matched_rows = 0
    unmatched_rows = 0
    ambiguous_rows = 0

    _log(logger, f"开始数据匹配：目标表 {options.target_path.name} <- 来源表 {options.source_path.name}")
    _log(logger, f"目标表表头位于第 {target_header_row} 行，来源表表头位于第 {source_header_row} 行。")
    _log(logger, f"主匹配列：目标[{options.target_key_column}] -> 来源[{options.source_key_column}]")
    if options.extra_match_mappings:
        _log(
            logger,
            "附加匹配列：" + "，".join(
                f"目标[{mapping.target_column}] -> 来源[{mapping.source_column}]"
                for mapping in options.extra_match_mappings
            ),
        )
    _log(
        logger,
        "补充列：" + "，".join(
            f"目标新增[{mapping.target_column or mapping.source_column}] <- 来源[{mapping.source_column}]"
            for mapping in options.transfer_mappings
        ),
    )

    for row_number, row in enumerate(target_rows, start=2):
        match_key = _make_key(
            row,
            target_index,
            options.target_key_column,
            options.extra_match_mappings,
            use_target_side=True,
        )
        key_text = " | ".join(match_key)
        if not match_key[0]:
            unmatched_rows += 1
            output_sheet.append(row + [""] * len(options.transfer_mappings))
            report_sheet.append([row_number, key_text, "未匹配", "主匹配列为空"])
            continue

        source_row = source_map.get(match_key)
        if match_key in duplicate_keys or source_row == []:
            ambiguous_rows += 1
            output_sheet.append(row + [""] * len(options.transfer_mappings))
            report_sheet.append([row_number, key_text, "未写入", "来源表存在重复匹配键"])
            continue
        if source_row is None:
            unmatched_rows += 1
            output_sheet.append(row + [""] * len(options.transfer_mappings))
            report_sheet.append([row_number, key_text, "未匹配", "来源表未找到对应记录"])
            continue

        append_values = [
            source_row[source_index[mapping.source_column]] for mapping in options.transfer_mappings
        ]
        output_sheet.append(row + append_values)
        matched_rows += 1
        report_sheet.append([row_number, key_text, "已匹配", ""])

    output_path = options.output_path or _build_output_path(options.target_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(output_path)
    _log(logger, f"数据匹配完成，结果已保存：{output_path}")

    return DataMatchSummary(
        target_path=options.target_path,
        source_path=options.source_path,
        output_path=output_path,
        target_key_column=options.target_key_column,
        source_key_column=options.source_key_column,
        extra_match_mappings=options.extra_match_mappings,
        transfer_mappings=options.transfer_mappings,
        total_rows=total_rows,
        matched_rows=matched_rows,
        unmatched_rows=unmatched_rows,
        duplicate_source_keys=len(duplicate_keys),
        ambiguous_rows=ambiguous_rows,
        target_header_row=target_header_row,
        source_header_row=source_header_row,
    )
