import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Optional


TEMPLATE_FILENAME = "照片分类模板.xlsx"
HEADERS = ["文件名称", "文件名前缀", "文件格式", "分类一", "分类二", "分类三", "修改名称"]


@dataclass
class TemplateResult:
    template_path: Path
    created: bool
    file_count: int


@dataclass
class ClassificationRecord:
    original_name: str
    destination_relative_path: str
    category_1: str
    category_2: str
    category_3: str
    renamed_name: str
    status: str
    note: str


@dataclass
class ClassificationResult:
    processed_count: int
    report_path: Optional[Path]
    records: List[ClassificationRecord]


def _log(logger: Optional[Callable[[str], None]], message: str) -> None:
    if logger is not None:
        logger(message)


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError(
            "缺少依赖 openpyxl，请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return Workbook, load_workbook


def list_current_level_files(directory: Path) -> List[Path]:
    files: List[Path] = []
    for path in sorted(directory.iterdir(), key=lambda item: item.name):
        if not path.is_file():
            continue
        if path.name == TEMPLATE_FILENAME:
            continue
        if path.name.startswith("."):
            continue
        files.append(path)
    return files


def split_filename_parts(filename: str) -> List[str]:
    path = Path(filename)
    return [path.stem, path.suffix]


def normalize_existing_values(row: tuple) -> List[str]:
    # 兼容旧模板列结构，避免历史模板在升级后直接错位。
    values = ["" if value is None else str(value) for value in row[1:]]
    if len(values) >= 6:
        return values[:6]
    if len(values) >= 4:
        return ["", "", values[0], values[1], values[2], values[3]]
    while len(values) < 6:
        values.append("")
    return values[:6]


def generate_template(
    source_dir: Path,
    dry_run: bool = False,
    logger: Optional[Callable[[str], None]] = None,
) -> TemplateResult:
    template_path = source_dir / TEMPLATE_FILENAME
    current_files = list_current_level_files(source_dir)
    existing_rows: Dict[str, List[str]] = {}
    created = not template_path.exists()
    Workbook, load_workbook = _load_openpyxl()

    if template_path.exists():
        workbook = load_workbook(template_path)
        worksheet = workbook.worksheets[0]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            filename = str(row[0]).strip()
            existing_rows[filename] = normalize_existing_values(row)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "分类模板"

    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(HEADERS)
    for file_path in current_files:
        prefix, suffix = split_filename_parts(file_path.name)
        values = existing_rows.get(file_path.name, ["", "", "", "", "", ""])
        # 前两列由程序重新拆分生成，后面的分类信息尽量沿用旧模板里已经填好的值。
        row_values = [file_path.name, prefix, suffix, *values[2:]]
        if values[0]:
            row_values[1] = values[0]
        if values[1]:
            row_values[2] = values[1]
        worksheet.append(row_values)

    if dry_run:
        _log(logger, f"[DRY-RUN] 将生成或更新 Excel 模板：{template_path}")
    else:
        workbook.save(template_path)
        if created:
            _log(logger, f"已生成 Excel 模板：{template_path}")
        else:
            _log(logger, f"已更新 Excel 模板：{template_path}")

    return TemplateResult(
        template_path=template_path,
        created=created,
        file_count=len(current_files),
    )


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        return ""
    return str(value).strip()


def _ensure_filename(original_name: str, renamed_value: str) -> str:
    if not renamed_value:
        return original_name
    candidate = renamed_value.strip()
    if not Path(candidate).suffix:
        candidate = candidate + Path(original_name).suffix
    return candidate


def _ensure_unique_path(destination: Path) -> Path:
    if not destination.exists():
        return destination

    stem = destination.stem
    suffix = destination.suffix
    index = 1
    while True:
        candidate = destination.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def validate_template_ready(template_path: Path) -> None:
    _, load_workbook = _load_openpyxl()
    workbook_formula = load_workbook(template_path, data_only=False)
    workbook_values = load_workbook(template_path, data_only=True)
    sheet_formula = workbook_formula.worksheets[0]
    sheet_values = workbook_values.worksheets[0]

    check_columns = [4, 5, 6, 7]
    # 只检查用户会写公式的业务列，避免把未计算的公式文本直接拿去做分类。
    for row_index, (formula_row, value_row) in enumerate(
        zip(
            sheet_formula.iter_rows(min_row=2),
            sheet_values.iter_rows(min_row=2, values_only=True),
        ),
        start=2,
    ):
        if not value_row:
            continue
        original_name = _normalize_cell(value_row[0] if len(value_row) > 0 else "")
        if not original_name:
            continue

        for col_index in check_columns:
            formula_cell = formula_row[col_index - 1]
            formula_value = formula_cell.value
            data_value = value_row[col_index - 1] if len(value_row) >= col_index else None
            if isinstance(formula_value, str) and formula_value.startswith("="):
                if _normalize_cell(data_value) == "":
                    raise ValueError(
                        f"检测到第一个 sheet 第 {row_index} 行存在尚未计算完成的公式。"
                        "请先在 Excel/WPS 中完成公式计算并保存，再执行分类。"
                    )


def export_classification_report(target_dir: Path, records: List[ClassificationRecord]) -> Path:
    Workbook, _ = _load_openpyxl()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "分类结果清单"
    worksheet.append(
        [
            "原文件名",
            "分类一",
            "分类二",
            "分类三",
            "修改名称",
            "输出相对路径",
            "状态",
            "备注",
        ]
    )
    for record in records:
        worksheet.append(
            [
                record.original_name,
                record.category_1,
                record.category_2,
                record.category_3,
                record.renamed_name,
                record.destination_relative_path,
                record.status,
                record.note,
            ]
        )
    report_path = target_dir / "照片分类结果清单.xlsx"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)
    return report_path


def apply_classification_from_template(
    source_dir: Path,
    target_dir: Path,
    dry_run: bool = False,
    logger: Optional[Callable[[str], None]] = None,
) -> ClassificationResult:
    template_path = source_dir / TEMPLATE_FILENAME
    if not template_path.exists():
        raise FileNotFoundError(f"Excel 模板不存在：{template_path}")

    validate_template_ready(template_path)
    _, load_workbook = _load_openpyxl()
    workbook = load_workbook(template_path, data_only=True)
    worksheet = workbook.worksheets[0]

    processed_count = 0
    records: List[ClassificationRecord] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        original_name = _normalize_cell(row[0] if len(row) > 0 else "")
        if not original_name:
            continue

        source_file = source_dir / original_name
        if not source_file.exists() or not source_file.is_file():
            _log(logger, f"[SKIP] 文件不存在：{source_file}")
            records.append(
                ClassificationRecord(
                    original_name=original_name,
                    destination_relative_path="",
                    category_1=_normalize_cell(row[3] if len(row) > 3 else ""),
                    category_2=_normalize_cell(row[4] if len(row) > 4 else ""),
                    category_3=_normalize_cell(row[5] if len(row) > 5 else ""),
                    renamed_name=_normalize_cell(row[6] if len(row) > 6 else ""),
                    status="跳过",
                    note="源文件不存在",
                )
            )
            continue

        categories = [
            _normalize_cell(row[3] if len(row) > 3 else ""),
            _normalize_cell(row[4] if len(row) > 4 else ""),
            _normalize_cell(row[5] if len(row) > 5 else ""),
        ]
        renamed_value = _normalize_cell(row[6] if len(row) > 6 else "")
        destination_name = _ensure_filename(original_name, renamed_value)

        destination_dir = target_dir
        for category in categories:
            if category:
                destination_dir = destination_dir / category

        # 根目录允许保留“未分类”文件；已分类文件只会进入具体分类目录。
        destination_path = _ensure_unique_path(destination_dir / destination_name)
        _log(logger, f"[COPY] {source_file} -> {destination_path}")
        if not dry_run:
            destination_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_file, destination_path)
        records.append(
            ClassificationRecord(
                original_name=original_name,
                destination_relative_path=str(destination_path.relative_to(target_dir)),
                category_1=categories[0],
                category_2=categories[1],
                category_3=categories[2],
                renamed_name=destination_name,
                status="预览" if dry_run else "已复制",
                note="",
            )
        )
        processed_count += 1

    _log(logger, f"分类复制完成，共处理 {processed_count} 个文件。")
    report_path = None
    if not dry_run:
        report_path = export_classification_report(target_dir, records)
        _log(logger, f"已导出结果清单：{report_path}")
    return ClassificationResult(
        processed_count=processed_count,
        report_path=report_path,
        records=records,
    )
