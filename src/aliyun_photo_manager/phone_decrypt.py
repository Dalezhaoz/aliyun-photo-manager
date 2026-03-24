from __future__ import annotations

import os
import json
import struct
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, List, Optional, Sequence


LogFn = Optional[Callable[[str], None]]


@dataclass
class PhoneDecryptOptions:
    server: str
    port: int
    username: str
    password: str
    signup_database: str
    phone_database: str
    candidate_table: str
    candidate_filter_mode: str = "all"
    candidate_id_cards: Optional[Sequence[str]] = None
    output_path: Optional[Path] = None


@dataclass
class PhoneDecryptRecord:
    primary_key: str
    id_card: str
    province: str
    encrypted_phone: str
    decrypted_phone: str
    status: str
    note: str


@dataclass
class PhoneDecryptSummary:
    signup_database: str
    phone_database: str
    candidate_table: str
    output_path: Path
    total_rows: int
    matched_info_rows: int
    decrypted_rows: int
    updated_rows: int
    skipped_rows: int
    failed_rows: int
    backend_name: str
    records: List[PhoneDecryptRecord]


@dataclass
class _PhoneDecryptRequest:
    primary_key: str
    encrypted_phone: str
    sort_code: str
    exam_date: str
    province: str


@dataclass
class _PhoneDecryptResponse:
    primary_key: str
    encrypted_phone: str
    decrypted_phone: str
    was_encrypted: bool
    success: bool
    error: str


def _log(logger: LogFn, message: str) -> None:
    if logger is not None:
        logger(message)


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_pyodbc():
    try:
        import pyodbc  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "未安装 pyodbc，无法连接 SQL Server。请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return pyodbc


def _ordered_drivers(pyodbc) -> list[str]:
    preferred = [
        "ODBC Driver 17 for SQL Server",
        "ODBC Driver 18 for SQL Server",
        "SQL Server",
    ]
    installed = list(pyodbc.drivers())
    ordered = [name for name in preferred if name in installed]
    # 仅追加名称中包含 "SQL Server" 的驱动，过滤掉 Access/Excel 等无关驱动
    for name in installed:
        if name not in ordered and "sql server" in name.lower():
            ordered.append(name)
    if not ordered:
        raise RuntimeError(
            "未找到 SQL Server ODBC 驱动。"
            " 请先安装 Microsoft ODBC Driver 17 for SQL Server："
            " https://learn.microsoft.com/zh-cn/sql/connect/odbc/download-odbc-driver-for-sql-server"
        )
    return ordered


def _probe_connection_string(connection_string: str) -> tuple[bool, str]:
    """在子进程中测试连接串是否可用，避免 ODBC 驱动 access violation 崩溃主进程。"""
    probe_code = (
        "import sys, pyodbc\n"
        "try:\n"
        "    conn = pyodbc.connect(sys.argv[1], timeout=6)\n"
        "    conn.close()\n"
        "    print('OK')\n"
        "except Exception as e:\n"
        "    print('ERR:' + str(e))\n"
        "    sys.exit(1)\n"
    )
    run_kwargs: dict = dict(
        capture_output=True,
        text=True,
        check=False,
        stdin=subprocess.DEVNULL,
        timeout=15,
    )
    if sys.platform == "win32":
        run_kwargs["creationflags"] = getattr(
            subprocess, "CREATE_NO_WINDOW", 0x08000000
        )
    try:
        result = subprocess.run(
            [sys.executable, "-c", probe_code, connection_string],
            **run_kwargs,
        )
    except subprocess.TimeoutExpired:
        return False, "连接探测超时。"
    except Exception as exc:
        return False, str(exc)
    if result.returncode == 0 and result.stdout.strip().startswith("OK"):
        return True, ""
    error = result.stdout.strip() or result.stderr.strip() or f"返回码 {result.returncode}"
    if error.startswith("ERR:"):
        error = error[4:]
    return False, error


class _DbConnection:
    """统一 pyodbc 和 pymssql 的接口差异。"""

    def __init__(self, raw_conn, backend: str) -> None:
        self._conn = raw_conn
        self.backend = backend
        # pymssql 用 %s，pyodbc 用 ?
        self._placeholder = "%s" if backend == "pymssql" else "?"

    def cursor(self):
        return self._conn.cursor()

    def commit(self):
        self._conn.commit()

    def close(self):
        self._conn.close()

    def rewrite_sql(self, sql: str) -> str:
        """将 SQL 中的 ? 占位符替换为当前后端使用的占位符。"""
        if self._placeholder == "?":
            return sql
        return sql.replace("?", self._placeholder)

    def set_fast_executemany(self, cursor) -> None:
        """仅 pyodbc 支持 fast_executemany。"""
        if self.backend == "pyodbc":
            cursor.fast_executemany = True


def _connect_pymssql(options: PhoneDecryptOptions, logger: LogFn = None) -> _DbConnection | None:
    """使用 pymssql 连接 SQL Server（不依赖 ODBC 驱动）。"""
    try:
        import pymssql  # type: ignore
    except ImportError:
        return None
    _log(logger, "尝试使用 pymssql 连接...")
    try:
        conn = pymssql.connect(
            server=options.server,
            port=options.port,
            user=options.username,
            password=options.password,
            login_timeout=8,
        )
        _log(logger, "pymssql 连接成功。")
        return _DbConnection(conn, "pymssql")
    except Exception as exc:
        _log(logger, f"pymssql 连接失败：{exc}")
        return None


def _connect_sql_server(options: PhoneDecryptOptions, logger: LogFn = None):
    # 优先尝试 pymssql（不依赖 ODBC，兼容性好）
    conn = _connect_pymssql(options, logger=logger)
    if conn is not None:
        return conn

    # 回退到 pyodbc
    pyodbc = _load_pyodbc()
    errors: List[str] = []
    for driver in _ordered_drivers(pyodbc):
        for encrypt_option in ("yes", "optional", "no"):
            connection_string = (
                f"DRIVER={{{driver}}};"
                f"SERVER={options.server},{options.port};"
                f"UID={options.username};"
                f"PWD={options.password};"
                "TrustServerCertificate=yes;"
                f"Encrypt={encrypt_option};"
            )
            _log(logger, f"尝试连接：{driver}  Encrypt={encrypt_option}")
            ok, err = _probe_connection_string(connection_string)
            if not ok:
                detail = f"[{driver}] Encrypt={encrypt_option}: {err}"
                _log(logger, f"  探测失败：{err}")
                errors.append(detail)
                continue
            _log(logger, f"  探测成功，正式连接...")
            try:
                return _DbConnection(pyodbc.connect(connection_string, timeout=8), "pyodbc")
            except Exception as exc:  # pragma: no cover
                detail = f"[{driver}] Encrypt={encrypt_option}: {exc}"
                _log(logger, f"  连接失败：{exc}")
                errors.append(detail)
    raise RuntimeError(
        "连接 SQL Server 失败，所有驱动均不可用。\n"
        "建议执行 pip install pymssql 安装替代驱动。\n"
        "详细错误：\n" + "\n".join(errors)
    )


def _escape_identifier(name: str) -> str:
    cleaned = name.strip()
    if not cleaned:
        raise ValueError("数据库名或表名不能为空。")
    return cleaned.replace("]", "]]")


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError("缺少依赖 openpyxl，请先执行 `pip install -r requirements.txt`。") from exc
    return Workbook, load_workbook


def _load_xlrd():
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError("缺少依赖 xlrd，请先执行 `pip install -r requirements.txt`。") from exc
    return xlrd


def _read_sheet_matrix(file_path: Path) -> List[List[str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        _, load_workbook = _load_openpyxl()
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        return [[_normalize(value) for value in row] for row in worksheet.iter_rows(values_only=True)]
    if suffix == ".xls":
        xlrd = _load_xlrd()
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        return [
            [_normalize(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
    raise ValueError("名单文件仅支持 `.xlsx` 或 `.xls`。")


def _trim_row(row: List[str]) -> List[str]:
    trimmed = list(row)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def _detect_header_row(matrix: List[List[str]]) -> int:
    best_index = 0
    best_score = -1
    for row_index in range(min(len(matrix), 10)):
        row = _trim_row(matrix[row_index])
        non_empty = [cell for cell in row if cell]
        if not non_empty:
            continue
        score = len(non_empty) * 10 + len(set(non_empty))
        if len(non_empty) == 1:
            score -= 20
        if score > best_score:
            best_score = score
            best_index = row_index
    return best_index


def load_filter_id_cards(file_path: Path) -> List[str]:
    if not file_path.exists():
        raise FileNotFoundError(f"未找到名单文件：{file_path}")
    matrix = _read_sheet_matrix(file_path)
    if not matrix:
        raise ValueError("名单文件为空。")
    header_index = _detect_header_row(matrix)
    headers = [header or f"未命名列{index + 1}" for index, header in enumerate(_trim_row(matrix[header_index]))]
    candidates = {"身份证号", "证件号码", "身份证", "sfzh"}
    target_index = next((index for index, header in enumerate(headers) if header.strip() in candidates), None)
    if target_index is None:
        raise ValueError("名单文件缺少“身份证号”列。")
    id_cards: List[str] = []
    seen = set()
    width = len(headers)
    for raw_row in matrix[header_index + 1 :]:
        row = list(raw_row[:width])
        if len(row) < width:
            row.extend([""] * (width - len(row)))
        id_card = _normalize(row[target_index])
        if id_card and id_card not in seen:
            id_cards.append(id_card)
            seen.add(id_card)
    return id_cards


class _DecryptorAdapter:
    def __init__(self, inner, backend_name: str) -> None:
        self.inner = inner
        self.backend_name = backend_name

    def check_encrypted(self, value: str) -> bool:
        return bool(self.inner.CheckEncrypted(value))

    def decrypt(self, data: str, sort_code: str, exam_date: str, province: str) -> str:
        return _normalize(self.inner.Decrypt(data, sort_code, exam_date, province))


def _candidate_component_dirs() -> List[Path]:
    dirs: List[Path] = []
    raw_candidates = [
        Path.cwd(),
        Path(__file__).resolve().parent,
        Path(sys.executable).resolve().parent,
    ]
    meipass = getattr(sys, "_MEIPASS", "")
    if meipass:
        raw_candidates.append(Path(meipass))
    env_dir = os.environ.get("PHONE_DECRYPT_DLL_DIR", "").strip()
    if env_dir:
        raw_candidates.append(Path(env_dir))

    seen = set()
    for item in raw_candidates:
        normalized = item.expanduser().resolve()
        if normalized in seen:
            continue
        seen.add(normalized)
        dirs.append(normalized)
    return dirs


def _find_component_path(filename: str) -> Path | None:
    for base_dir in _candidate_component_dirs():
        candidate = base_dir / filename
        if candidate.exists():
            return candidate
    return None


def _candidate_helper_paths() -> List[Path]:
    candidates: List[Path] = []
    env_path = os.environ.get("PHONE_DECRYPT_HELPER", "").strip()
    if env_path:
        candidates.append(Path(env_path))

    meipass = getattr(sys, "_MEIPASS", "")
    if meipass:
        candidates.append(Path(meipass) / "PhoneDecryptHelper.exe")

    project_root = Path(__file__).resolve().parents[2]
    candidates.extend(
        [
            project_root / "PhoneDecryptHelper.exe",
            project_root / "tools" / "PhoneDecryptHelper" / "bin" / "Release" / "net8.0-windows" / "win-x86" / "publish" / "PhoneDecryptHelper.exe",
            project_root / "tools" / "PhoneDecryptHelper" / "bin" / "Release" / "net8.0-windows" / "PhoneDecryptHelper.exe",
            Path.cwd() / "PhoneDecryptHelper.exe",
            Path(sys.executable).resolve().parent / "PhoneDecryptHelper.exe",
        ]
    )
    seen = set()
    result: List[Path] = []
    for item in candidates:
        resolved = item.expanduser().resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        result.append(resolved)
    return result


def _find_helper_path() -> Path | None:
    for candidate in _candidate_helper_paths():
        if candidate.exists():
            return candidate
    return None


def _run_helper_batch(requests: Sequence[_PhoneDecryptRequest], logger: LogFn = None) -> tuple[str, dict[str, _PhoneDecryptResponse]]:
    helper_path = _find_helper_path()
    if helper_path is None:
        searched = "，".join(str(item) for item in _candidate_helper_paths())
        raise RuntimeError(
            "未找到 PhoneDecryptHelper.exe。"
            " 请先在 Windows 上编译 32 位 helper，或设置环境变量 PHONE_DECRYPT_HELPER。"
            f" 已搜索：{searched}"
        )

    payload = {
        "requests": [
            {
                "primaryKey": item.primary_key,
                "encryptedPhone": item.encrypted_phone,
                "sortCode": item.sort_code,
                "examDate": item.exam_date,
                "province": item.province,
            }
            for item in requests
        ]
    }
    with tempfile.TemporaryDirectory(prefix="phone_decrypt_") as temp_dir:
        temp_path = Path(temp_dir)
        input_path = temp_path / "input.json"
        output_path = temp_path / "output.json"
        input_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        _log(logger, f"调用 32 位电话解密 helper：{helper_path}")
        run_kwargs: dict = dict(
            capture_output=True,
            text=True,
            check=False,
            stdin=subprocess.DEVNULL,
        )
        if sys.platform == "win32":
            run_kwargs["creationflags"] = getattr(
                subprocess, "CREATE_NO_WINDOW", 0x08000000
            )
        result = subprocess.run(
            [str(helper_path), str(input_path), str(output_path)],
            **run_kwargs,
        )
        if result.returncode != 0:
            details = [f"helper 返回码: {result.returncode}"]
            if result.stderr.strip():
                details.append(f"stderr: {result.stderr.strip()}")
            if result.stdout.strip():
                details.append(f"stdout: {result.stdout.strip()}")
            raise RuntimeError("电话解密 helper 执行失败\n" + "\n".join(details))
        if not output_path.exists():
            raise RuntimeError("电话解密 helper 未生成输出文件。")
        data = json.loads(output_path.read_text(encoding="utf-8"))
    responses = {
        str(item.get("primaryKey", "")): _PhoneDecryptResponse(
            primary_key=str(item.get("primaryKey", "")),
            encrypted_phone=str(item.get("encryptedPhone", "")),
            decrypted_phone=str(item.get("decryptedPhone", "")),
            was_encrypted=bool(item.get("wasEncrypted", False)),
            success=bool(item.get("success", False)),
            error=str(item.get("error", "")),
        )
        for item in data.get("responses", [])
        if isinstance(item, dict)
    }
    backend_name = str(data.get("backendName", helper_path.name)).strip() or helper_path.name
    return backend_name, responses


def _load_pythonnet_decryptor() -> _DecryptorAdapter:
    try:
        import clr  # type: ignore
    except ImportError as exc:
        raise RuntimeError("未安装 pythonnet，无法加载电话解密组件。") from exc

    interop_path = _find_component_path("Interop.DeDll.dll")
    core_dll_path = _find_component_path("DeDLL.dll")
    errors: List[str] = []

    if interop_path is not None:
        try:
            if core_dll_path is not None:
                os.environ["PATH"] = str(core_dll_path.parent) + os.pathsep + os.environ.get("PATH", "")
            clr.AddReference(str(interop_path))
            import DeDll  # type: ignore

            return _DecryptorAdapter(DeDll.DeAESClass(), f"pythonnet:{interop_path.name}")
        except Exception as exc:  # pragma: no cover
            errors.append(f"按文件路径加载 {interop_path} 失败：{exc}")
    else:
        searched = "，".join(str(item) for item in _candidate_component_dirs())
        errors.append(
            "未找到 Interop.DeDll.dll。"
            f" 请把 `Interop.DeDll.dll` 和 `DeDLL.dll` 放到程序目录或设置环境变量 `PHONE_DECRYPT_DLL_DIR`。已搜索：{searched}"
        )

    for assembly_name in ("Interop.DeDll", "DeDll"):
        try:
            clr.AddReference(assembly_name)
            import DeDll  # type: ignore

            return _DecryptorAdapter(DeDll.DeAESClass(), f"pythonnet:{assembly_name}")
        except Exception as exc:  # pragma: no cover
            errors.append(f"按程序集名加载 {assembly_name} 失败：{exc}")
    raise RuntimeError("；".join(errors))


def _load_win32com_decryptor() -> _DecryptorAdapter:
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError("未安装 pywin32，无法通过 COM 调用电话解密组件。") from exc

    errors: List[str] = []
    for prog_id in (
        "DeDll.DeAES",
        "DeDll.DeAESClass",
        "DeDLL.DeAES",
        "DeDLL.DeAESClass",
    ):
        try:
            return _DecryptorAdapter(win32com.client.Dispatch(prog_id), f"win32com:{prog_id}")
        except Exception as exc:  # pragma: no cover
            errors.append(f"{prog_id} -> {exc}")
    raise RuntimeError("创建 DeDll 解密对象失败：" + "；".join(errors))


def load_phone_decryptor() -> _DecryptorAdapter:
    if os.name != "nt":
        raise RuntimeError("电话解密仅支持 Windows 环境。")

    python_bits = struct.calcsize("P") * 8
    if python_bits != 32:
        raise RuntimeError(
            f"当前 Python 为 {python_bits} 位，无法直接加载 32 位解密 DLL。"
            " 请确保 PhoneDecryptHelper.exe（32 位 helper）可用，"
            " 或使用 32 位 Python 运行本程序。"
        )

    errors: List[str] = []
    for loader in (_load_pythonnet_decryptor, _load_win32com_decryptor):
        try:
            return loader()
        except Exception as exc:
            errors.append(str(exc))
    if errors:
        raise RuntimeError("；".join(errors))
    raise RuntimeError("未能加载电话解密组件。")


def _normalize_province(raw_province: str, sort_code: str) -> str:
    province = _normalize(raw_province)
    if province.startswith("141") and sort_code == "96":
        return "141"
    return province


def _split_primary_key(primary_key: str) -> tuple[str, str]:
    normalized = _normalize(primary_key)
    if len(normalized) < 8:
        raise ValueError(f"主键编号长度不足，无法拆分考试代码和考试年月：{normalized}")
    return normalized[:2], normalized[2:8]


def _iter_chunked(items: Sequence[str], size: int = 500) -> Iterable[Sequence[str]]:
    for index in range(0, len(items), size):
        yield items[index : index + size]


def _fetch_candidate_rows(cursor, connection: _DbConnection, options: PhoneDecryptOptions) -> List[tuple[str, str, str, str]]:
    signup_db = _escape_identifier(options.signup_database)
    phone_db = _escape_identifier(options.phone_database or options.signup_database)
    candidate_table = _escape_identifier(options.candidate_table)
    base_sql = f"""
    SELECT
        LTRIM(RTRIM(CAST(ks.[主键编号] AS VARCHAR(50)))) AS primary_key,
        LTRIM(RTRIM(ISNULL(CAST(ks.[身份证号] AS VARCHAR(50)), ''))) AS id_card,
        LTRIM(RTRIM(ISNULL(CAST(ks.[考区] AS VARCHAR(50)), ''))) AS province,
        LTRIM(RTRIM(ISNULL(CAST(info.[info1] AS VARCHAR(255)), ''))) AS encrypted_phone
    FROM [{signup_db}].[dbo].[{candidate_table}] ks
    LEFT JOIN [{phone_db}].[dbo].[web_info] info
        ON CAST(info.[zjbh] AS VARCHAR(50)) = CAST(ks.[主键编号] AS VARCHAR(50))
       AND CAST(info.[examsort] AS VARCHAR(10)) = LEFT(CAST(ks.[主键编号] AS VARCHAR(50)), 2)
    WHERE ks.[主键编号] IS NOT NULL
    """

    if options.candidate_filter_mode != "partial" or not options.candidate_id_cards:
        cursor.execute(base_sql + " ORDER BY ks.[主键编号]")
        return [(str(row[0]), str(row[1]), str(row[2]), str(row[3])) for row in cursor.fetchall()]

    placeholder = connection._placeholder
    rows: List[tuple[str, str, str, str]] = []
    for chunk in _iter_chunked(list(options.candidate_id_cards)):
        placeholders = ",".join(placeholder for _ in chunk)
        cursor.execute(
            base_sql + f" AND CAST(ks.[身份证号] AS VARCHAR(50)) IN ({placeholders}) ORDER BY ks.[主键编号]",
            list(chunk),
        )
        rows.extend((str(row[0]), str(row[1]), str(row[2]), str(row[3])) for row in cursor.fetchall())
    return rows


def _export_phone_report(records: List[PhoneDecryptRecord], output_path: Path) -> None:
    Workbook, _ = _load_openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "电话解密结果"
    sheet.append(["主键编号", "身份证号", "考区", "加密串", "解密后电话", "状态", "备注"])
    for item in records:
        sheet.append(
            [
                item.primary_key,
                item.id_card,
                item.province,
                item.encrypted_phone,
                item.decrypted_phone,
                item.status,
                item.note,
            ]
        )
    workbook.save(output_path)


def _build_output_path(options: PhoneDecryptOptions) -> Path:
    if options.output_path is not None:
        return options.output_path
    return Path.cwd() / f"{options.candidate_table}_电话解密结果.xlsx"


def run_phone_decrypt(options: PhoneDecryptOptions, logger: LogFn = None) -> PhoneDecryptSummary:
    if not options.server.strip():
        raise ValueError("请输入服务器地址。")
    if not options.username.strip() or not options.password.strip():
        raise ValueError("请输入数据库用户名和密码。")
    if not options.signup_database.strip():
        raise ValueError("请输入报名数据库名称。")
    if not options.candidate_table.strip():
        raise ValueError("请输入考生表名称。")

    output_path = _build_output_path(options)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _log(logger, f"开始查询考生表：{options.signup_database}.{options.candidate_table}")

    connection = _connect_sql_server(options, logger=logger)
    try:
        cursor = connection.cursor()
        rows = _fetch_candidate_rows(cursor, connection, options)
        _log(logger, f"共读取到 {len(rows)} 条待处理记录。")

        records: List[PhoneDecryptRecord] = []
        updates: List[tuple[str, str]] = []
        matched_info_rows = 0
        decrypted_rows = 0
        skipped_rows = 0
        failed_rows = 0
        backend_name = ""
        pending_requests: List[_PhoneDecryptRequest] = []
        pending_rows: dict[str, tuple[str, str, str]] = {}

        for primary_key, id_card, province, encrypted_phone in rows:
            encrypted_phone = _normalize(encrypted_phone)
            if encrypted_phone:
                matched_info_rows += 1
            if not encrypted_phone:
                skipped_rows += 1
                records.append(
                    PhoneDecryptRecord(
                        primary_key=primary_key,
                        id_card=id_card,
                        province=province,
                        encrypted_phone="",
                        decrypted_phone="",
                        status="跳过",
                        note="未找到电话密文（web_info.info1 为空）。",
                    )
                )
                continue

            try:
                sort_code, exam_date = _split_primary_key(primary_key)
                province_value = _normalize_province(province, sort_code)
                if not province_value:
                    raise ValueError("考区代码为空。")
            except Exception as exc:
                failed_rows += 1
                records.append(
                    PhoneDecryptRecord(
                        primary_key=primary_key,
                        id_card=id_card,
                        province=province,
                        encrypted_phone=encrypted_phone,
                        decrypted_phone="",
                        status="失败",
                        note=str(exc),
                    )
                )
                continue
            pending_requests.append(
                _PhoneDecryptRequest(
                    primary_key=primary_key,
                    encrypted_phone=encrypted_phone,
                    sort_code=sort_code,
                    exam_date=exam_date,
                    province=province_value,
                )
            )
            pending_rows[primary_key] = (id_card, province, encrypted_phone)

        responses: dict[str, _PhoneDecryptResponse] = {}
        if pending_requests:
            try:
                backend_name, responses = _run_helper_batch(pending_requests, logger=logger)
            except Exception as helper_exc:
                _log(logger, f"32 位 helper 不可用，尝试进程内加载：{helper_exc}")
                decryptor = load_phone_decryptor()
                backend_name = decryptor.backend_name
                _log(logger, f"已加载电话解密组件：{backend_name}")
                for request in pending_requests:
                    try:
                        if decryptor.check_encrypted(request.encrypted_phone):
                            decrypted_phone = decryptor.decrypt(
                                request.encrypted_phone,
                                request.sort_code,
                                request.exam_date,
                                request.province,
                            )
                            was_encrypted = True
                        else:
                            decrypted_phone = request.encrypted_phone
                            was_encrypted = False
                        responses[request.primary_key] = _PhoneDecryptResponse(
                            primary_key=request.primary_key,
                            encrypted_phone=request.encrypted_phone,
                            decrypted_phone=decrypted_phone,
                            was_encrypted=was_encrypted,
                            success=bool(decrypted_phone),
                            error="" if decrypted_phone else "解密结果为空。",
                        )
                    except Exception as exc:
                        responses[request.primary_key] = _PhoneDecryptResponse(
                            primary_key=request.primary_key,
                            encrypted_phone=request.encrypted_phone,
                            decrypted_phone="",
                            was_encrypted=True,
                            success=False,
                            error=str(exc),
                        )
            else:
                _log(logger, f"已加载电话解密组件：{backend_name}")

        for request in pending_requests:
            id_card, province, encrypted_phone = pending_rows[request.primary_key]
            response = responses.get(request.primary_key)
            if response is None:
                failed_rows += 1
                records.append(
                    PhoneDecryptRecord(
                        primary_key=request.primary_key,
                        id_card=id_card,
                        province=province,
                        encrypted_phone=encrypted_phone,
                        decrypted_phone="",
                        status="失败",
                        note="未收到解密结果。",
                    )
                )
                continue
            if not response.success:
                failed_rows += 1
                records.append(
                    PhoneDecryptRecord(
                        primary_key=request.primary_key,
                        id_card=id_card,
                        province=province,
                        encrypted_phone=encrypted_phone,
                        decrypted_phone="",
                        status="失败",
                        note=response.error or "解密失败。",
                    )
                )
                continue
            decrypted_phone = _normalize(response.decrypted_phone)
            decrypted_rows += 1
            updates.append((decrypted_phone, request.primary_key))
            records.append(
                PhoneDecryptRecord(
                    primary_key=request.primary_key,
                    id_card=id_card,
                    province=province,
                    encrypted_phone=encrypted_phone,
                    decrypted_phone=decrypted_phone,
                    status="成功",
                    note="已调用 DLL 解密。" if response.was_encrypted else "原值未加密，直接写回。",
                )
            )

        if updates:
            signup_db = _escape_identifier(options.signup_database)
            candidate_table = _escape_identifier(options.candidate_table)
            update_sql = connection.rewrite_sql(f"""
            UPDATE [{signup_db}].[dbo].[{candidate_table}]
            SET [备用3] = ?
            WHERE CAST([主键编号] AS VARCHAR(50)) = ?
            """)
            connection.set_fast_executemany(cursor)
            cursor.executemany(update_sql, updates)
            connection.commit()
            _log(logger, f"已更新 {len(updates)} 条电话到备用3。")
        else:
            _log(logger, "没有可回写的电话记录。")

        _export_phone_report(records, output_path)
        _log(logger, f"已导出电话解密结果清单：{output_path}")

        return PhoneDecryptSummary(
            signup_database=options.signup_database,
            phone_database=options.phone_database or options.signup_database,
            candidate_table=options.candidate_table,
            output_path=output_path,
            total_rows=len(rows),
            matched_info_rows=matched_info_rows,
            decrypted_rows=decrypted_rows,
            updated_rows=len(updates),
            skipped_rows=skipped_rows,
            failed_rows=failed_rows,
            backend_name=backend_name or "unknown",
            records=records,
        )
    finally:
        connection.close()
