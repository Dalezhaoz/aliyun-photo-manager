import shutil
from dataclasses import dataclass
from pathlib import Path
from threading import Event
from typing import Callable, Dict, List, Optional

from .downloader import DownloadResult


LogFn = Optional[Callable[[str], None]]
ProgressFn = Optional[Callable[[str, int, int, str], None]]


@dataclass
class CertificateFilterOptions:
    template_path: Path
    source_dir: Path
    output_dir: Path
    match_column: str
    rename_folder: bool = False
    folder_name_column: str = ""
    classify_output: bool = False
    keyword: str = ""
    dry_run: bool = False


@dataclass
class CertificateFilterSummary:
    template_path: Path
    source_dir: Path
    output_dir: Path
    match_column: str
    rename_folder: bool
    folder_name_column: str
    classify_output: bool
    keyword: str
    total_rows: int
    matched_people: int
    missing_people: int
    copied_files: int
    copied_people: int
    download_result: Optional[DownloadResult] = None
    cancelled: bool = False
    dry_run: bool = False
    report_path: Optional[Path] = None


@dataclass
class CertificateRecord:
    match_value: str
    category_1: str
    category_2: str
    category_3: str
    output_relative_dir: str
    copied_files: int
    status: str
    note: str


def _log(logger: LogFn, message: str) -> None:
    if logger is not None:
        logger(message)


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError(
            "缺少依赖 openpyxl，请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return Workbook, load_workbook


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def list_template_headers(template_path: Path) -> List[str]:
    _, load_workbook = _load_openpyxl()
    workbook = load_workbook(template_path, data_only=True)
    worksheet = workbook.worksheets[0]
    headers: List[str] = []
    for cell in worksheet[1]:
        header = _normalize_cell(cell.value)
        if header:
            headers.append(header)
    return headers


def _read_template_rows(template_path: Path) -> List[Dict[str, str]]:
    _, load_workbook = _load_openpyxl()
    workbook = load_workbook(template_path, data_only=True)
    worksheet = workbook.worksheets[0]

    headers = [_normalize_cell(cell.value) for cell in worksheet[1]]
    rows: List[Dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        item: Dict[str, str] = {}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = _normalize_cell(row[index] if len(row) > index else "")
            item[header] = value
            if value:
                has_value = True
        if has_value:
            rows.append(item)
    return rows


def _iter_files(directory: Path) -> List[Path]:
    return [path for path in directory.rglob("*") if path.is_file()]


def _copy_person_folder(
    source_dir: Path,
    destination_dir: Path,
    keyword: str,
    dry_run: bool,
) -> int:
    copied_files = 0
    normalized_keyword = keyword.strip().lower()

    for source_file in _iter_files(source_dir):
        if normalized_keyword and normalized_keyword not in source_file.name.lower():
            continue
        relative_path = source_file.relative_to(source_dir)
        destination_path = destination_dir / relative_path
        copied_files += 1
        if dry_run:
            continue
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_file, destination_path)

    return copied_files


def _export_certificate_report(output_dir: Path, records: List[CertificateRecord]) -> Path:
    Workbook, _ = _load_openpyxl()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "筛选结果清单"
    worksheet.append(
        [
            "匹配值",
            "分类一",
            "分类二",
            "分类三",
            "输出相对目录",
            "复制文件数",
            "状态",
            "备注",
        ]
    )
    for record in records:
        worksheet.append(
            [
                record.match_value,
                record.category_1,
                record.category_2,
                record.category_3,
                record.output_relative_dir,
                record.copied_files,
                record.status,
                record.note,
            ]
        )
    report_path = output_dir / "证件资料筛选结果清单.xlsx"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)
    return report_path


def run_certificate_filter(
    options: CertificateFilterOptions,
    logger: LogFn = None,
    progress_callback: ProgressFn = None,
    cancel_event: Optional[Event] = None,
) -> CertificateFilterSummary:
    if not options.template_path.exists():
        raise FileNotFoundError(f"模板文件不存在：{options.template_path}")
    if not options.source_dir.exists():
        raise FileNotFoundError(f"证件资料目录不存在：{options.source_dir}")

    rows = _read_template_rows(options.template_path)
    headers = list(rows[0].keys()) if rows else list_template_headers(options.template_path)
    if options.match_column not in headers:
        raise ValueError(f"模板中不存在匹配列：{options.match_column}")
    if options.rename_folder and options.folder_name_column not in headers:
        raise ValueError(f"模板中不存在导出后文件夹名称列：{options.folder_name_column}")

    total_rows = len(rows)
    matched_people = 0
    missing_people = 0
    copied_files = 0
    copied_people = 0
    seen_matches = set()
    records: List[CertificateRecord] = []

    if progress_callback is not None:
        progress_callback("certificate", 0, total_rows, "")

    mode_label = "整个人员文件夹" if not options.keyword.strip() else f"关键词文件（{options.keyword.strip()}）"
    _log(logger, f"开始证件资料筛选。模式：{mode_label}")
    _log(logger, f"匹配列：{options.match_column}")

    for index, row in enumerate(rows, start=1):
        if cancel_event is not None and cancel_event.is_set():
            _log(logger, f"证件资料筛选已取消，已处理 {index - 1}/{total_rows} 人。")
            return CertificateFilterSummary(
                template_path=options.template_path,
                source_dir=options.source_dir,
                output_dir=options.output_dir,
                match_column=options.match_column,
                classify_output=options.classify_output,
                keyword=options.keyword.strip(),
                total_rows=total_rows,
                matched_people=matched_people,
                missing_people=missing_people,
                copied_files=copied_files,
                copied_people=copied_people,
                download_result=None,
                cancelled=True,
                dry_run=options.dry_run,
                report_path=None,
            )

        match_value = row.get(options.match_column, "").strip()
        if not match_value:
            if progress_callback is not None:
                progress_callback("certificate", index, total_rows, "")
            continue
        if match_value in seen_matches:
            _log(logger, f"[SKIP] 模板中重复匹配值，已跳过：{match_value}")
            records.append(
                CertificateRecord(
                    match_value=match_value,
                    category_1=row.get("分类一", "").strip(),
                    category_2=row.get("分类二", "").strip(),
                    category_3=row.get("分类三", "").strip(),
                    output_relative_dir="",
                    copied_files=0,
                    status="跳过",
                    note="模板中重复匹配值",
                )
            )
            if progress_callback is not None:
                progress_callback("certificate", index, total_rows, match_value)
            continue
        seen_matches.add(match_value)

        matched_people += 1
        person_source_dir = options.source_dir / match_value
        if not person_source_dir.exists() or not person_source_dir.is_dir():
            missing_people += 1
            _log(logger, f"[SKIP] 未找到人员文件夹：{person_source_dir}")
            records.append(
                CertificateRecord(
                    match_value=match_value,
                    category_1=row.get("分类一", "").strip(),
                    category_2=row.get("分类二", "").strip(),
                    category_3=row.get("分类三", "").strip(),
                    output_relative_dir="",
                    copied_files=0,
                    status="缺失",
                    note="未找到人员文件夹",
                )
            )
            if progress_callback is not None:
                progress_callback("certificate", index, total_rows, match_value)
            continue

        output_folder_name = match_value
        if options.rename_folder:
            output_folder_name = row.get(options.folder_name_column, "").strip() or match_value

        destination_dir = options.output_dir
        if options.classify_output:
            for field in ("分类一", "分类二", "分类三"):
                value = row.get(field, "").strip()
                if value:
                    destination_dir = destination_dir / value
        destination_dir = destination_dir / output_folder_name
        relative_destination_dir = str(destination_dir.relative_to(options.output_dir))

        current_file_count = _copy_person_folder(
            source_dir=person_source_dir,
            destination_dir=destination_dir,
            keyword=options.keyword,
            dry_run=options.dry_run,
        )
        if current_file_count > 0:
            copied_people += 1
            copied_files += current_file_count
            records.append(
                CertificateRecord(
                    match_value=match_value,
                    category_1=row.get("分类一", "").strip(),
                    category_2=row.get("分类二", "").strip(),
                    category_3=row.get("分类三", "").strip(),
                    output_relative_dir=relative_destination_dir,
                    copied_files=current_file_count,
                    status="预览" if options.dry_run else "已复制",
                    note="",
                )
            )
            if index == 1 or index % 50 == 0 or index == total_rows:
                _log(
                    logger,
                    f"证件资料进度 {index}/{total_rows}：{match_value}，复制 {current_file_count} 个文件。",
                )
        else:
            _log(logger, f"[SKIP] {match_value} 下没有匹配文件。")
            records.append(
                CertificateRecord(
                    match_value=match_value,
                    category_1=row.get("分类一", "").strip(),
                    category_2=row.get("分类二", "").strip(),
                    category_3=row.get("分类三", "").strip(),
                    output_relative_dir=relative_destination_dir,
                    copied_files=0,
                    status="跳过",
                    note="没有匹配文件",
                )
            )

        if progress_callback is not None:
            progress_callback("certificate", index, total_rows, match_value)

    _log(
        logger,
        f"证件资料筛选完成：匹配 {matched_people} 人，"
        f"缺失 {missing_people} 人，复制 {copied_people} 人的 {copied_files} 个文件。",
    )
    report_path = None
    if not options.dry_run:
        report_path = _export_certificate_report(options.output_dir, records)
        _log(logger, f"已导出结果清单：{report_path}")
    return CertificateFilterSummary(
        template_path=options.template_path,
        source_dir=options.source_dir,
        output_dir=options.output_dir,
        match_column=options.match_column,
        rename_folder=options.rename_folder,
        folder_name_column=options.folder_name_column,
        classify_output=options.classify_output,
        keyword=options.keyword.strip(),
        total_rows=total_rows,
        matched_people=matched_people,
        missing_people=missing_people,
        copied_files=copied_files,
        copied_people=copied_people,
        download_result=None,
        cancelled=False,
        dry_run=options.dry_run,
        report_path=report_path,
    )
