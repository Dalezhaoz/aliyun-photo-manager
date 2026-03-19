from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, List, Optional, Tuple


LogFn = Optional[Callable[[str], None]]


@dataclass
class UpdateSqlFieldMapping:
    target_column: str
    source_column: str
    enabled: bool


@dataclass
class UpdateSqlResult:
    mapping_path: Path
    target_table: str
    source_table: str
    target_key_column: str
    source_key_column: str
    ignore_empty: bool
    backup_target_table: str
    backup_source_table: str
    updated_columns: List[Tuple[str, str]]
    sql_content: str


@dataclass
class UpdateSqlTemplateExportSummary:
    output_path: Path


def _log(logger: LogFn, message: str) -> None:
    if logger is not None:
        logger(message)


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError("缺少依赖 openpyxl，请先执行 `pip install -r requirements.txt`。") from exc
    return Workbook, load_workbook


def _load_xlrd():
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError("缺少依赖 xlrd，请先执行 `pip install -r requirements.txt`。") from exc
    return xlrd


def _read_sheet_matrix(file_path: Path) -> List[List[str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        _, load_workbook = _load_openpyxl()
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        return [[_normalize(value) for value in row] for row in worksheet.iter_rows(values_only=True)]
    if suffix == ".xls":
        xlrd = _load_xlrd()
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        return [
            [_normalize(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
    raise ValueError("仅支持 `.xlsx` 或 `.xls` 文件。")


def _trim_row(row: List[str]) -> List[str]:
    trimmed = list(row)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def _detect_header_row(matrix: List[List[str]]) -> int:
    best_index = 0
    best_score = -1
    for row_index in range(min(len(matrix), 10)):
        row = _trim_row(matrix[row_index])
        non_empty = [cell for cell in row if cell]
        if not non_empty:
            continue
        score = len(non_empty) * 10 + len(set(non_empty))
        if len(non_empty) == 1:
            score -= 20
        if score > best_score:
            best_score = score
            best_index = row_index
    return best_index


def _extract_headers_and_rows(matrix: List[List[str]]) -> Tuple[List[str], List[List[str]]]:
    if not matrix:
        raise ValueError("模板为空，无法读取字段映射。")
    header_index = _detect_header_row(matrix)
    raw_headers = _trim_row(matrix[header_index])
    headers = [header or f"未命名列{index + 1}" for index, header in enumerate(raw_headers)]
    width = len(headers)
    rows: List[List[str]] = []
    for raw_row in matrix[header_index + 1 :]:
        row = list(raw_row[:width])
        if len(row) < width:
            row.extend([""] * (width - len(row)))
        row = [_normalize(value) for value in row]
        if any(row):
            rows.append(row)
    return headers, rows


def _quote_identifier(name: str) -> str:
    parts = [part.strip() for part in name.split(".") if part.strip()]
    if not parts:
        raise ValueError("表名或字段名不能为空。")
    return ".".join(f"[{part.replace(']', ']]')}]" for part in parts)


def load_update_field_mappings(mapping_path: Path) -> Tuple[List[UpdateSqlFieldMapping], List[str], List[str]]:
    if not mapping_path.exists():
        raise FileNotFoundError(f"未找到映射模板：{mapping_path}")

    headers, rows = _extract_headers_and_rows(_read_sheet_matrix(mapping_path))
    required_headers = {"考生表字段名", "临时表字段名", "是否更新"}
    missing = required_headers - set(headers)
    if missing:
        raise ValueError(f"映射模板缺少必要列：{', '.join(sorted(missing))}")

    target_index = headers.index("考生表字段名")
    source_index = headers.index("临时表字段名")
    enabled_index = headers.index("是否更新")

    mappings: List[UpdateSqlFieldMapping] = []
    target_values: List[str] = []
    source_values: List[str] = []
    seen_target = set()
    seen_source = set()

    for row in rows:
        target_column = _normalize(row[target_index])
        source_column = _normalize(row[source_index])
        enabled_text = _normalize(row[enabled_index]).lower()
        enabled = enabled_text in {"是", "y", "yes", "true", "1"}
        if target_column and target_column not in seen_target:
            target_values.append(target_column)
            seen_target.add(target_column)
        if source_column and source_column not in seen_source:
            source_values.append(source_column)
            seen_source.add(source_column)
        if target_column and source_column:
            mappings.append(
                UpdateSqlFieldMapping(
                    target_column=target_column,
                    source_column=source_column,
                    enabled=enabled,
                )
            )

    return mappings, target_values, source_values


def export_update_sql_template(output_path: Path) -> UpdateSqlTemplateExportSummary:
    Workbook, _ = _load_openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "字段映射"
    sheet.append(["考生表字段名", "临时表字段名", "是否更新"])
    sheet.append(["xm", "姓名", "是"])
    sheet.append(["sfzh", "身份证号", "否"])
    workbook.save(output_path)
    return UpdateSqlTemplateExportSummary(output_path=output_path)


def render_update_sql(
    mapping_path: Path,
    target_table: str,
    source_table: str,
    target_key_column: str,
    source_key_column: str,
    ignore_empty: bool,
    logger: LogFn = None,
) -> UpdateSqlResult:
    mappings, _, _ = load_update_field_mappings(mapping_path)
    effective_mappings = [
        mapping
        for mapping in mappings
        if mapping.enabled and mapping.target_column and mapping.source_column
    ]
    if not effective_mappings:
        raise ValueError("映射模板中没有可更新的字段，请把“是否更新”列设置为“是”。")
    if not target_table.strip() or not source_table.strip():
        raise ValueError("请输入考生表名称和临时表名称。")
    if not target_key_column.strip() or not source_key_column.strip():
        raise ValueError("请选择关联字段。")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_target_table = f"{target_table.strip()}_bak_{timestamp}"
    backup_source_table = f"{source_table.strip()}_bak_{timestamp}"

    _log(logger, f"读取字段映射模板：{mapping_path.name}")
    _log(logger, f"生成更新 SQL：{target_table} <- {source_table}")

    assignment_lines: List[str] = []
    updated_columns: List[Tuple[str, str]] = []
    for mapping in effective_mappings:
        target_expr = f"ks.{_quote_identifier(mapping.target_column)}"
        source_expr = f"tmp.{_quote_identifier(mapping.source_column)}"
        if ignore_empty:
            assignment = (
                f"{target_expr} = CASE "
                f"WHEN {source_expr} IS NOT NULL AND LTRIM(RTRIM(CONVERT(NVARCHAR(MAX), {source_expr}))) <> '' "
                f"THEN {source_expr} ELSE {target_expr} END"
            )
        else:
            assignment = f"{target_expr} = {source_expr}"
        assignment_lines.append(assignment)
        updated_columns.append((mapping.target_column, mapping.source_column))

    target_table_sql = _quote_identifier(target_table.strip())
    source_table_sql = _quote_identifier(source_table.strip())
    target_key_sql = _quote_identifier(target_key_column.strip())
    source_key_sql = _quote_identifier(source_key_column.strip())
    backup_target_sql = _quote_identifier(backup_target_table)
    backup_source_sql = _quote_identifier(backup_source_table)

    sql_lines = [
        "-- 生成时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "-- 说明：先备份正式表和临时表，再执行字段更新。",
        "BEGIN TRAN;",
        "",
        f"SELECT * INTO {backup_target_sql}",
        f"FROM {target_table_sql};",
        "",
        f"SELECT * INTO {backup_source_sql}",
        f"FROM {source_table_sql};",
        "",
        "UPDATE ks",
        "SET",
        "    " + ",\n    ".join(assignment_lines),
        f"FROM {target_table_sql} ks",
        f"INNER JOIN {source_table_sql} tmp",
        f"    ON ks.{target_key_sql} = tmp.{source_key_sql};",
        "",
        "COMMIT TRAN;",
    ]

    return UpdateSqlResult(
        mapping_path=mapping_path,
        target_table=target_table.strip(),
        source_table=source_table.strip(),
        target_key_column=target_key_column.strip(),
        source_key_column=source_key_column.strip(),
        ignore_empty=ignore_empty,
        backup_target_table=backup_target_table,
        backup_source_table=backup_source_table,
        updated_columns=updated_columns,
        sql_content="\n".join(sql_lines),
    )
