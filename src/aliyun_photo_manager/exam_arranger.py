from dataclasses import dataclass
from pathlib import Path
import random
from typing import Callable, Dict, List, Optional, Tuple


LogFn = Optional[Callable[[str], None]]

CANDIDATE_HEADERS = ["姓名", "身份证号", "招聘单位", "岗位名称"]
GROUP_HEADERS = ["招聘单位", "岗位名称", "科目组", "岗位编码"]
GROUP_TEMPLATE_HEADERS = ["招聘单位", "岗位名称", "科目组", "岗位编码", "科目号"]
PLAN_HEADERS = ["科目组", "考点", "考场号", "起始座号", "结束座号", "人数", "起始流水号", "结束流水号", "备注"]


@dataclass
class ExamRuleItem:
    item_type: str
    custom_text: str = ""


@dataclass
class ExamArrangeOptions:
    candidate_path: Path
    group_path: Path
    plan_path: Path
    output_path: Optional[Path]
    exam_point_digits: int
    room_digits: int
    seat_digits: int
    serial_digits: int
    sort_mode: str
    rule_items: List[ExamRuleItem]


@dataclass
class ExamArrangeSummary:
    candidate_path: Path
    group_path: Path
    plan_path: Path
    output_path: Path
    total_candidates: int
    arranged_candidates: int
    missing_groups: int
    missing_plan_groups: int
    duplicate_group_rows: int
    unused_plan_slots: int


@dataclass
class ExamTemplateExportSummary:
    output_dir: Path
    candidate_template_path: Path
    group_template_path: Path
    plan_template_path: Path


def _log(logger: LogFn, message: str) -> None:
    if logger is not None:
        logger(message)


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value, field_name: str) -> int:
    text = _normalize(value)
    if not text:
        raise ValueError(f"{field_name} 不能为空。")
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"{field_name} 必须是整数：{text}") from exc


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError(
            "缺少依赖 openpyxl，请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return Workbook, load_workbook


def _load_xlrd():
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(
            "缺少依赖 xlrd，请先执行 `pip install -r requirements.txt`。"
        ) from exc
    return xlrd


def _read_matrix(file_path: Path) -> List[List[str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        _, load_workbook = _load_openpyxl()
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        return [[_normalize(value) for value in row] for row in worksheet.iter_rows(values_only=True)]
    if suffix == ".xls":
        xlrd = _load_xlrd()
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        return [
            [_normalize(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
    raise ValueError("仅支持 `.xlsx` 或 `.xls` 文件。")


def _trim_row(row: List[str]) -> List[str]:
    values = list(row)
    while values and values[-1] == "":
        values.pop()
    return values


def _find_header_row(matrix: List[List[str]], required_headers: List[str]) -> int:
    required = set(required_headers)
    best_index = -1
    best_score = -1
    for row_index, raw_row in enumerate(matrix[:15]):
        row = set(cell for cell in _trim_row(raw_row) if cell)
        score = len(required & row)
        if score > best_score:
            best_score = score
            best_index = row_index
    if best_score <= 0:
        raise ValueError(f"未识别到表头，请确认表中包含这些列：{', '.join(required_headers)}")
    return best_index


def _read_records(file_path: Path, required_headers: List[str]) -> List[Dict[str, str]]:
    matrix = _read_matrix(file_path)
    if not matrix:
        raise ValueError(f"文件为空：{file_path}")
    header_row_index = _find_header_row(matrix, required_headers)
    header_row = _trim_row(matrix[header_row_index])
    header_map = {header: index for index, header in enumerate(header_row) if header}
    missing = [header for header in required_headers if header not in header_map]
    if missing:
        raise ValueError(f"{file_path.name} 缺少必要列：{', '.join(missing)}")

    records: List[Dict[str, str]] = []
    width = len(header_row)
    for row in matrix[header_row_index + 1 :]:
        normalized = list(row[:width])
        if len(normalized) < width:
            normalized.extend([""] * (width - len(normalized)))
        normalized = [_normalize(value) for value in normalized]
        if not any(normalized):
            continue
        record = {header: normalized[index] if index < len(normalized) else "" for header, index in header_map.items()}
        records.append(record)
    return records


def list_headers(file_path: Path, required_headers: Optional[List[str]] = None) -> List[str]:
    matrix = _read_matrix(file_path)
    if not matrix:
        raise ValueError(f"文件为空：{file_path}")
    reference_headers = required_headers or []
    header_row_index = _find_header_row(matrix, reference_headers) if reference_headers else 0
    header_row = _trim_row(matrix[header_row_index])
    headers = [header for header in header_row if header]
    if required_headers:
        missing = [header for header in required_headers if header not in headers]
        if missing:
            raise ValueError(f"{file_path.name} 缺少必要列：{', '.join(missing)}")
    return headers


def _default_output_path(candidate_path: Path) -> Path:
    return candidate_path.with_name(f"{candidate_path.stem}_考场编排结果.xlsx")


def export_exam_templates(output_dir: Path) -> ExamTemplateExportSummary:
    """Export the three standard exam arrangement templates for manual completion."""
    output_dir.mkdir(parents=True, exist_ok=True)
    Workbook, _ = _load_openpyxl()

    candidate_template_path = output_dir / "考生明细模板.xlsx"
    candidate_workbook = Workbook()
    candidate_sheet = candidate_workbook.active
    candidate_sheet.title = "考生明细"
    candidate_sheet.append(CANDIDATE_HEADERS)
    candidate_sheet.append(["张三", "370101199001011234", "某招聘单位", "某岗位名称"])
    candidate_workbook.save(candidate_template_path)

    group_template_path = output_dir / "岗位归组模板.xlsx"
    group_workbook = Workbook()
    group_sheet = group_workbook.active
    group_sheet.title = "岗位归组"
    group_sheet.append(GROUP_TEMPLATE_HEADERS)
    group_sheet.append(["某招聘单位", "某岗位名称", "语文组", "01", "02"])
    group_workbook.save(group_template_path)

    plan_template_path = output_dir / "编排片段模板.xlsx"
    plan_workbook = Workbook()
    plan_sheet = plan_workbook.active
    plan_sheet.title = "编排片段"
    plan_sheet.append(PLAN_HEADERS)
    # 一行描述一个考场片段，混编时同一考场可以拆成多行分别描述不同座位区间。
    plan_sheet.append(["语文组", "01", "001", "1", "30", "30", "1", "30", "普通整场"])
    plan_sheet.append(["语文组", "01", "009", "1", "5", "5", "151", "155", "尾数混编"])
    plan_workbook.save(plan_template_path)

    return ExamTemplateExportSummary(
        output_dir=output_dir,
        candidate_template_path=candidate_template_path,
        group_template_path=group_template_path,
        plan_template_path=plan_template_path,
    )


def _format_number(value: str, digits: int) -> str:
    text = _normalize(value)
    if not text:
        return ""
    try:
        numeric = int(float(text))
        return str(numeric).zfill(digits)
    except ValueError:
        return text.zfill(digits)


def _build_exam_number(
    item_values: Dict[str, str],
    rule_items: List[ExamRuleItem],
    exam_point_digits: int,
    room_digits: int,
    seat_digits: int,
    serial_digits: int,
) -> str:
    parts: List[str] = []
    for rule in rule_items:
        if rule.item_type == "自定义":
            parts.append(rule.custom_text.strip())
        elif rule.item_type == "考点":
            parts.append(_format_number(item_values.get("考点", ""), exam_point_digits))
        elif rule.item_type == "考场":
            parts.append(_format_number(item_values.get("考场", ""), room_digits))
        elif rule.item_type == "座号":
            parts.append(_format_number(item_values.get("座号", ""), seat_digits))
        elif rule.item_type == "流水号":
            parts.append(_format_number(item_values.get("流水号", ""), serial_digits))
        elif rule.item_type in item_values:
            parts.append(item_values.get(rule.item_type, "").strip())
    return "".join(parts)


def run_exam_arrangement(options: ExamArrangeOptions, logger: LogFn = None) -> ExamArrangeSummary:
    if not options.rule_items:
        raise ValueError("请至少配置一条考号规则。")
    candidate_records = _read_records(options.candidate_path, CANDIDATE_HEADERS)
    group_records = _read_records(options.group_path, GROUP_HEADERS)
    plan_records = _read_records(options.plan_path, PLAN_HEADERS)

    group_map: Dict[Tuple[str, str], Dict[str, str]] = {}
    duplicate_group_rows = 0
    for record in group_records:
        key = (record.get("招聘单位", ""), record.get("岗位名称", ""))
        if key in group_map:
            duplicate_group_rows += 1
            continue
        group_map[key] = record

    plan_map: Dict[str, List[Dict[str, str]]] = {}
    for record in plan_records:
        group_name = record.get("科目组", "")
        if not group_name:
            continue
        plan_map.setdefault(group_name, []).append(record)

    grouped_candidates: Dict[str, List[Dict[str, str]]] = {}
    missing_groups = 0
    unresolved_candidates: List[Dict[str, str]] = []
    for record in candidate_records:
        key = (record.get("招聘单位", ""), record.get("岗位名称", ""))
        group_record = group_map.get(key)
        if group_record is None:
            missing_groups += 1
            unresolved_candidates.append(record)
            continue
        enriched = dict(record)
        enriched["科目组"] = group_record.get("科目组", "")
        enriched["岗位编码"] = group_record.get("岗位编码", "")
        enriched["科目号"] = group_record.get("科目号", "")
        grouped_candidates.setdefault(enriched["科目组"], []).append(enriched)

    arranged_rows: List[Dict[str, str]] = []
    missing_plan_groups = 0
    unused_plan_slots = 0
    arranged_candidates = 0

    _log(logger, f"开始考场编排：共 {len(candidate_records)} 人。")
    _log(logger, f"岗位归组表共 {len(group_records)} 行，编排片段表共 {len(plan_records)} 行。")
    _log(logger, f"同组内排序方式：{'随机打乱' if options.sort_mode == 'random' else '按原顺序'}。")

    for group_name, candidates in grouped_candidates.items():
        if options.sort_mode == "random":
            # 仅在组内打乱，不改变跨组分配规则。
            random.shuffle(candidates)
        segments = plan_map.get(group_name, [])
        if not segments:
            missing_plan_groups += len(candidates)
            for candidate in candidates:
                row = dict(candidate)
                row.update({"考点": "", "考场": "", "座号": "", "考号": "", "编排备注": "未找到科目组编排片段"})
                arranged_rows.append(row)
            continue

        candidate_index = 0
        for segment in segments:
            start_seat = _to_int(segment.get("起始座号", ""), "起始座号")
            end_seat = _to_int(segment.get("结束座号", ""), "结束座号")
            start_serial = _to_int(segment.get("起始流水号", ""), "起始流水号")
            end_serial = _to_int(segment.get("结束流水号", ""), "结束流水号")
            seat_capacity = end_seat - start_seat + 1
            serial_capacity = end_serial - start_serial + 1
            count_value = _to_int(segment.get("人数", ""), "人数")
            capacity = min(seat_capacity, serial_capacity, count_value)
            for offset in range(capacity):
                if candidate_index >= len(candidates):
                    unused_plan_slots += capacity - offset
                    break
                candidate = dict(candidates[candidate_index])
                candidate["考点"] = segment.get("考点", "")
                candidate["考场"] = segment.get("考场号", "")
                candidate["座号"] = str(start_seat + offset)
                candidate["流水号"] = str(start_serial + offset)
                candidate["编排备注"] = segment.get("备注", "")
                candidate["考号"] = _build_exam_number(
                    candidate,
                    options.rule_items,
                    options.exam_point_digits,
                    options.room_digits,
                    options.seat_digits,
                    options.serial_digits,
                )
                arranged_rows.append(candidate)
                arranged_candidates += 1
                candidate_index += 1
            if candidate_index >= len(candidates):
                remaining_segments = segments[segments.index(segment) + 1 :]
                for remain in remaining_segments:
                    unused_plan_slots += _to_int(remain.get("人数", ""), "人数")
                break

        if candidate_index < len(candidates):
            for candidate in candidates[candidate_index:]:
                row = dict(candidate)
                row.update({"考点": "", "考场": "", "座号": "", "考号": "", "编排备注": "编排片段人数不足"})
                arranged_rows.append(row)

    for record in unresolved_candidates:
        row = dict(record)
        row.update({"科目组": "", "岗位编码": "", "科目号": "", "考点": "", "考场": "", "座号": "", "考号": "", "编排备注": "未找到岗位归组"})
        arranged_rows.append(row)

    output_path = options.output_path or _default_output_path(options.candidate_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    Workbook, _ = _load_openpyxl()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "考场编排结果"
    worksheet.append(CANDIDATE_HEADERS + ["科目组", "岗位编码", "科目号", "考点", "考场", "座号", "考号", "编排备注"])
    for row in arranged_rows:
        worksheet.append(
            [
                row.get("姓名", ""),
                row.get("身份证号", ""),
                row.get("招聘单位", ""),
                row.get("岗位名称", ""),
                row.get("科目组", ""),
                row.get("岗位编码", ""),
                row.get("科目号", ""),
                row.get("考点", ""),
                row.get("考场", ""),
                row.get("座号", ""),
                row.get("考号", ""),
                row.get("编排备注", ""),
            ]
        )
    workbook.save(output_path)
    _log(logger, f"考场编排完成，结果已保存：{output_path}")

    return ExamArrangeSummary(
        candidate_path=options.candidate_path,
        group_path=options.group_path,
        plan_path=options.plan_path,
        output_path=output_path,
        total_candidates=len(candidate_records),
        arranged_candidates=arranged_candidates,
        missing_groups=missing_groups,
        missing_plan_groups=missing_plan_groups,
        duplicate_group_rows=duplicate_group_rows,
        unused_plan_slots=unused_plan_slots,
    )
